"""
build_panel.py
==============
Constructs a tidy annual panel dataset shaped for energy poverty transition
matrix analysis:

    Year | Decile | Energy_Spend_Proxy | Income_EUR | Unemployment_Rate
         | Elec_CPI_YoY | Gas_CPI_YoY | State

Two temporal windows are stacked:

  A) 2007–2016 : no decile-level inflation data exists from CSO.
     Deciles 1–10 are SIMULATED by applying an income-share scaling factor
     to the aggregate housing CPI YoY.  Income is approximated by linear
     interpolation of the SIA13 age-group medians mapped to decile rank.
     All "simulated" rows are flagged: Data_Source = "interpolated".

  B) 2017–2023 : CSO EIHC01 provides actual monthly YoY inflation per decile.
     Annual average of monthly readings used.
     Income per decile is estimated by mapping decile rank : income percentile
     on the SIA13 age-group curve (18–64 median as D5 anchor).
     All real rows are flagged: Data_Source = "observed".

Energy Poverty State Classification
------------------------------------
A decile-year is classified as "Energy Poor" if BOTH:
  1. Energy_Spend_Proxy > national_median_energy_spend * 2   (2× median rule)
  2. Decile rank ≤ 4  (bottom 40% of income distribution)

"At Risk" if only condition 1 holds (high energy burden but not lowest income).
"Not Poor" otherwise.

Energy_Spend_Proxy = Energy_CPI_composite_YoY × Income_share_of_energy
  where Income_share_of_energy is assumed 10% for D1, declining 0.5pp per
  decile to 5.5% for D10 (consistent with HBS 2015/2019/2022 averages for
  Ireland; Sustainable Energy Authority of Ireland estimates).

Output
------
  Cleaned Data/Panel_Energy_Poverty.csv    : full long-format panel
  Cleaned Data/Panel_Energy_Poverty_Annotated.xlsx: Excel with colour coding
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT     = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "Cleaned Data"
OUT_DIR  = ROOT / "Output"
OUT_DIR.mkdir(exist_ok=True)
OUT_CSV  = OUT_DIR / "Panel_Energy_Poverty.csv"
OUT_XLSX = OUT_DIR / "Panel_Energy_Poverty_Annotated.xlsx"

# ── 1. LOAD SOURCE DATA ───────────────────────────────────────────────────────

housing_cpi = pd.read_csv(DATA_DIR / "CPI_Housing_2007_2023.csv")
decile_inf  = pd.read_csv(DATA_DIR / "Inflation_by_Income_Decile_2017_2023.csv")
energy_cpi  = pd.read_csv(DATA_DIR / "CPI_Energy_Products_2007_2023.csv")
income_age  = pd.read_csv(DATA_DIR / "Income_Median_by_AgeGroup_2007_2023.csv")
unemp_raw   = pd.read_csv(DATA_DIR / "Unemployment_Rate_Monthly_2007_2023.csv")

# ── 2. ANNUAL AGGREGATES ──────────────────────────────────────────────────────

# 2a. Housing CPI: annual mean of monthly YoY %
housing_annual = (
    housing_cpi
    .groupby("Year")["YoY_Pct_Change"]
    .mean()
    .round(2)
    .reset_index()
    .rename(columns={"YoY_Pct_Change": "Housing_CPI_YoY"})
)

# 2b. Energy CPI: pivot to wide, compute YoY, keep electricity + gas
energy_wide = energy_cpi.pivot(
    index="Year", columns="Energy_Product", values="CPI_Index"
)
energy_yoy = energy_wide.pct_change() * 100
energy_annual = energy_yoy[["Electricity", "Natural gas"]].rename(
    columns={"Electricity": "Elec_CPI_YoY", "Natural gas": "Gas_CPI_YoY"}
).round(2).reset_index()
# Composite energy CPI = equal-weight average of electricity + gas YoY
energy_annual["Energy_CPI_Composite_YoY"] = (
    energy_annual[["Elec_CPI_YoY", "Gas_CPI_YoY"]].mean(axis=1).round(2)
)

# 2c. Unemployment: annual mean of monthly SA rate
unemp_raw["Year"] = unemp_raw["Month"].str[:4].astype(int)
unemp_annual = (
    unemp_raw
    .groupby("Year")["Unemployment_Rate_Pct"]
    .mean()
    .round(2)
    .reset_index()
    .rename(columns={"Unemployment_Rate_Pct": "Unemployment_Rate_Pct"})
)

# 2d. Income: 18-64 median as the working-age anchor (best proxy for D5)
# SIA13 data only available up to 2019; extrapolate 2020-2023 using a
# simple linear trend from the last 5 observed years (2015–2019).
income_64 = (
    income_age
    .loc[income_age["Age_Group"] == "18 - 64 years",
         ["Year", "Median_Real_Disposable_Income_EUR"]]
    .rename(columns={"Median_Real_Disposable_Income_EUR": "Income_D5_Anchor_EUR"})
    .sort_values("Year")
)
# Linear extrapolation from 2015-2019 trend
_trend_base = income_64[income_64["Year"] >= 2015].copy()
if len(_trend_base) >= 2:
    _coef = np.polyfit(_trend_base["Year"], _trend_base["Income_D5_Anchor_EUR"], 1)
    _poly = np.poly1d(_coef)
    _last_obs = int(income_64["Year"].max())
    _extrap = pd.DataFrame({
        "Year": range(_last_obs + 1, 2024),
        "Income_D5_Anchor_EUR": [round(_poly(y), 0) for y in range(_last_obs + 1, 2024)],
    })
    income_anchor = pd.concat([income_64, _extrap], ignore_index=True)
else:
    income_anchor = income_64

# 2e. Decile inflation: annual mean of monthly readings (observed 2017–2023)
decile_inf["Year"] = decile_inf["Month_dt"].str[:4].astype(int)
decile_annual_obs = (
    decile_inf[decile_inf["Income_Decile"] != "All deciles"]
    .groupby(["Year", "Income_Decile"])["YoY_Inflation_Pct"]
    .mean()
    .round(2)
    .reset_index()
)

# Standardise decile label : integer rank 1–10
decile_map = {
    "1st decile": 1, "2nd decile": 2,  "3rd decile": 3,
    "4th decile": 4, "5th decile": 5,  "6th decile": 6,
    "7th decile": 7, "8th decile": 8,  "9th decile": 9,
    "10th decile": 10,
}
decile_annual_obs["Decile"] = (
    decile_annual_obs["Income_Decile"].map(decile_map)
)
decile_annual_obs = decile_annual_obs.rename(
    columns={"YoY_Inflation_Pct": "Inflation_YoY_Pct"}
)[["Year", "Decile", "Inflation_YoY_Pct"]]

# ── 3. INCOME PER DECILE ──────────────────────────────────────────────────────
# Strategy: use the 18-64 median as D5 anchor, then apply a scaling curve
# based on typical Irish income distribution shape (Gini ≈ 0.31).
# Relative income by decile (D5 = 1.0) estimated from SILC 2015–2022 averages:
DECILE_INCOME_SCALE = {
    1: 0.28,   # D1 ≈ 28% of median
    2: 0.47,
    3: 0.60,
    4: 0.73,
    5: 0.87,
    6: 1.00,   # D5/D6 boundary ≈ median
    7: 1.15,
    8: 1.33,
    9: 1.58,
    10: 2.21,  # D10 ≈ 2.2× median (top decile)
}

# Assumed household energy expenditure share by decile (% of disposable income)
# Based on SEAI/HBS evidence: lower deciles spend disproportionately more on energy
ENERGY_SHARE = {
    1: 0.108,   # 10.8%
    2: 0.095,
    3: 0.085,
    4: 0.077,
    5: 0.071,
    6: 0.066,
    7: 0.062,
    8: 0.059,
    9: 0.057,
    10: 0.055,  # 5.5%
}

# ── 4. BUILD FULL PANEL (all years × all deciles) ─────────────────────────────

YEARS   = list(range(2007, 2024))
DECILES = list(range(1, 11))

rows = []
for year in YEARS:
    # Fetch scalar contextual values for this year
    h_row   = housing_annual.loc[housing_annual["Year"] == year]
    e_row   = energy_annual.loc[energy_annual["Year"] == year]
    u_row   = unemp_annual.loc[unemp_annual["Year"] == year]
    inc_row = income_anchor.loc[income_anchor["Year"] == year]

    housing_yoy  = float(h_row["Housing_CPI_YoY"].iloc[0]) if len(h_row) else np.nan
    elec_yoy     = float(e_row["Elec_CPI_YoY"].iloc[0])    if len(e_row) else np.nan
    gas_yoy      = float(e_row["Gas_CPI_YoY"].iloc[0])     if len(e_row) else np.nan
    composite_yoy= float(e_row["Energy_CPI_Composite_YoY"].iloc[0]) if len(e_row) else housing_yoy
    unemp_rate   = float(u_row["Unemployment_Rate_Pct"].iloc[0]) if len(u_row) else np.nan
    d5_income    = float(inc_row["Income_D5_Anchor_EUR"].iloc[0]) if len(inc_row) else np.nan

    for decile in DECILES:
        # ── Inflation YoY for this decile-year ──────────────────────────────
        if year >= 2017:
            obs = decile_annual_obs.loc[
                (decile_annual_obs["Year"] == year) &
                (decile_annual_obs["Decile"] == decile),
                "Inflation_YoY_Pct"
            ]
            inf_yoy     = float(obs.iloc[0]) if len(obs) > 0 else np.nan
            data_source = "observed"
        else:
            # Pre-2017: scale aggregate housing CPI by a decile-specific
            # sensitivity factor (lower deciles face ~15-25% higher energy inflation)
            # Scaling: D1 = housing_yoy × 1.20, D10 = housing_yoy × 0.85
            sensitivity = 1.20 - (decile - 1) * (0.35 / 9)
            inf_yoy     = round(housing_yoy * sensitivity, 2) if not np.isnan(housing_yoy) else np.nan
            data_source = "interpolated"

        # ── Income for this decile-year ──────────────────────────────────────
        income_eur = round(d5_income * DECILE_INCOME_SCALE[decile]) if not np.isnan(d5_income) else np.nan

        # ── Energy spend proxy = income × energy share × (1 + composite_yoy/100) ─
        # i.e., estimated annual energy cost in real terms
        energy_share_pct = ENERGY_SHARE[decile]
        if not np.isnan(income_eur) and not np.isnan(composite_yoy):
            energy_spend_proxy = round(income_eur * energy_share_pct * (1 + composite_yoy / 100), 0)
        elif not np.isnan(income_eur):
            energy_spend_proxy = round(income_eur * energy_share_pct, 0)
        else:
            energy_spend_proxy = np.nan

        rows.append({
            "Year":                    year,
            "Decile":                  decile,
            "Income_EUR":              income_eur,
            "Inflation_YoY_Pct":       inf_yoy,
            "Elec_CPI_YoY":            round(elec_yoy, 2) if not np.isnan(elec_yoy) else np.nan,
            "Gas_CPI_YoY":             round(gas_yoy, 2)  if not np.isnan(gas_yoy)  else np.nan,
            "Energy_CPI_Composite_YoY":round(composite_yoy, 2) if not np.isnan(composite_yoy) else np.nan,
            "Energy_Spend_Proxy_EUR":  energy_spend_proxy,
            "Energy_Share_Pct":        round(energy_share_pct * 100, 1),
            "Unemployment_Rate_Pct":   unemp_rate,
            "Data_Source":             data_source,
        })

panel = pd.DataFrame(rows)

# ── 5. CLASSIFY ENERGY POVERTY STATE ─────────────────────────────────────────
# Standard Irish energy poverty definition (SEAI / DHLGH):
#   Energy Poor = household spends > 10% of net income on energy
# We operationalise this using the decile-specific energy share assumption.
# Since Energy_Share_Pct is the assumed share, we classify based on whether
# the REAL (inflation-adjusted) energy burden has crossed the 10% threshold.
#
# Real energy burden = Energy_Share_Pct × (1 + cumulative CPI uplift)
# A simpler proxy: if the base share for a decile + inflation drag pushes
# effective share above 10%, that decile-year is "Energy Poor".
#
# Effective_Energy_Burden_Pct = Energy_Share_Pct × (1 + Inflation_YoY_Pct/100)
#
# States:
#   Energy Poor  : effective burden ≥ 10% AND Decile ≤ 5
#   At Risk      : effective burden ≥ 10% (any decile) OR (Decile ≤ 3 AND burden ≥ 8%)
#   Not Poor     : otherwise
#
# National median energy spend is also retained for reference.

median_spend = (
    panel.groupby("Year")["Energy_Spend_Proxy_EUR"]
    .median()
    .rename("Median_Energy_Spend_EUR")
)
panel = panel.merge(median_spend, on="Year")

# Effective burden: base energy share scaled up by this year's YoY inflation
panel["Effective_Burden_Pct"] = (
    panel["Energy_Share_Pct"] * (1 + panel["Inflation_YoY_Pct"].fillna(0) / 100)
).round(2)

def classify(row):
    if pd.isna(row["Energy_Spend_Proxy_EUR"]):
        return np.nan
    burden = row["Effective_Burden_Pct"]
    decile = row["Decile"]
    if burden >= 10.0 and decile <= 5:
        return "Energy Poor"
    if burden >= 10.0 or (decile <= 3 and burden >= 8.0):
        return "At Risk"
    return "Not Poor"

panel["State"] = panel.apply(classify, axis=1)

# Clean column order
COLS = [
    "Year", "Decile",
    "Income_EUR", "Energy_Share_Pct", "Energy_Spend_Proxy_EUR",
    "Effective_Burden_Pct",
    "Inflation_YoY_Pct",
    "Elec_CPI_YoY", "Gas_CPI_YoY", "Energy_CPI_Composite_YoY",
    "Unemployment_Rate_Pct",
    "Median_Energy_Spend_EUR", "State", "Data_Source",
]
panel = panel[COLS].sort_values(["Year", "Decile"]).reset_index(drop=True)

# ── 6. SAVE CSV ───────────────────────────────────────────────────────────────
panel.to_csv(OUT_CSV, index=False)

# ── 7. ANNOTATED EXCEL ───────────────────────────────────────────────────────
STATE_COLOURS = {
    "Energy Poor": "C00000",   # red
    "At Risk":     "F4B942",   # amber
    "Not Poor":    "70AD47",   # green
}
SOURCE_FILL = {
    "observed":     "FFFFFF",
    "interpolated": "FFF2CC",  # light yellow: flag for reader
}

wb = Workbook()
ws = wb.active
ws.title = "Energy Poverty Panel"

thin = Side(style="thin", color="D9D9D9")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR  = Alignment(horizontal="center", vertical="center")
LFT  = Alignment(horizontal="left",   vertical="center")

# Header
for ci, col in enumerate(panel.columns, start=1):
    c = ws.cell(1, ci, col)
    c.font      = Font(bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor="1F4E79")
    c.alignment = CTR
    c.border    = bdr
ws.row_dimensions[1].height = 20

# Data rows
for ri, row_data in panel.iterrows():
    state  = row_data["State"]
    source = row_data["Data_Source"]
    bg     = SOURCE_FILL.get(source, "FFFFFF")

    for ci, val in enumerate(row_data, start=1):
        col_name = panel.columns[ci - 1]
        cell = ws.cell(ri + 2, ci, val if not (isinstance(val, float) and np.isnan(val)) else "")
        cell.border    = bdr
        cell.alignment = CTR if col_name != "State" else CTR

        if col_name == "State" and state in STATE_COLOURS:
            cell.fill = PatternFill("solid", fgColor=STATE_COLOURS[state])
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        else:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10)

# Column widths
COL_WIDTHS = {
    "Year": 6, "Decile": 7, "Income_EUR": 13, "Energy_Share_Pct": 13,
    "Energy_Spend_Proxy_EUR": 20, "Effective_Burden_Pct": 20,
    "Inflation_YoY_Pct": 16,
    "Elec_CPI_YoY": 13, "Gas_CPI_YoY": 12, "Energy_CPI_Composite_YoY": 22,
    "Unemployment_Rate_Pct": 20, "Median_Energy_Spend_EUR": 22,
    "State": 13, "Data_Source": 14,
}
for ci, col in enumerate(panel.columns, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)

ws.freeze_panes = "A2"

# Legend sheet
ls = wb.create_sheet("Legend")
legend_rows = [
    ("State", "Colour", "Definition"),
    ("Energy Poor", "Red",   "Effective energy burden ≥ 10% of income AND Decile ≤ 5"),
    ("At Risk",     "Amber", "Effective burden ≥ 10% (any decile) OR Decile ≤ 3 AND burden ≥ 8%"),
    ("Not Poor",    "Green", "Effective energy burden below threshold"),
    ("", "", ""),
    ("Data_Source", "Background", "Meaning"),
    ("observed",     "White",       "Actual CSO EIHC01 decile inflation (2017–2023)"),
    ("interpolated", "Light Yellow","Pre-2017: housing CPI scaled by decile sensitivity factor"),
    ("", "", ""),
    ("Effective_Burden_Pct",   "", "Energy_Share_Pct × (1 + Inflation_YoY_Pct/100)"),
    ("Energy_Spend_Proxy_EUR", "", "Income × Energy_Share_Pct × (1 + Energy_CPI_Composite_YoY/100)"),
    ("Energy_Share_Pct",       "", "Assumed household energy budget share: D1=10.8%, D10=5.5%"),
    ("Income_EUR",             "", "D5-anchor (SIA13 18-64 median) × decile scale; 2020-23 extrapolated"),
]
for ri, row in enumerate(legend_rows, start=1):
    for ci, val in enumerate(row, start=1):
        ls.cell(ri, ci, val).font = Font(size=10, bold=(ri == 1 or ri == 6))
for ci in [1, 2, 3]:
    ls.column_dimensions[get_column_letter(ci)].width = 28

wb.save(OUT_XLSX)

# ── 8. SUMMARY PRINT ─────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("  PANEL DATASET: Energy Poverty Transition Analysis")
print(f"{'='*65}")
print(f"  Shape          : {panel.shape[0]} rows × {panel.shape[1]} cols")
print(f"  Period         : {panel['Year'].min()}–{panel['Year'].max()}")
print(f"  Deciles        : {panel['Decile'].min()}–{panel['Decile'].max()}")
print(f"\n  Data coverage:")
print(f"    Observed  (2017–2023): {(panel['Data_Source']=='observed').sum()} rows")
print(f"    Interpolated (2007–2016): {(panel['Data_Source']=='interpolated').sum()} rows")
print(f"\n  State distribution:")
print(panel.groupby(["State", "Data_Source"]).size().to_string())
print(f"\n  State by year (count of deciles in each state):")
print(panel.pivot_table(index="Year", columns="State", values="Decile",
                        aggfunc="count", fill_value=0).to_string())
print(f"\n  Sample rows:")
print(panel[panel["Year"].isin([2007, 2017, 2022])].to_string(index=False))
print(f"\n✅  CSV   : {OUT_CSV}")
print(f"✅  Excel : {OUT_XLSX}")
