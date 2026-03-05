"""
panel_descriptive_stats.py
==========================
Produces submission-ready descriptive statistics for the Energy Poverty Panel
dataset.  Output: Cleaned Data/Panel_Descriptive_Statistics.xlsx

Sheet layout
------------
  1. "Table 1: Variable Summary"   Overall pooled stats (N, mean, sd, min, p25,
                                    median, p75, max, skewness) for all numeric
                                    variables.  Separately for observed vs full panel.
  2. "Table 2: By State"           Means ± SD of each numeric variable broken down
                                    by State (Energy Poor / At Risk / Not Poor).
  3. "Table 3: By Decile"          Mean of key vars by Decile (annual panel averages).
  4. "Table 4: Annual Trends"      Year-level panel aggregate: mean income, mean
                                    energy burden, % in each state, mean inflation.
  5. "Table 5: State Frequency"    Count and % of decile-years in each state,
                                    by year: the raw material for transition matrices.
  6. "Table 6: Correlation Matrix" Pearson correlations among numeric variables
                                    (observed rows only).
"""

import numpy as np
import pandas as pd
from pathlib import Path
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── PATHS ─────────────────────────────────────────────────────────────────────
ROOT     = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "Cleaned Data"
PANEL    = ROOT / "Output" / "Panel_Energy_Poverty.csv"
OUT_DIR  = ROOT / "Output"
OUT_DIR.mkdir(exist_ok=True)
OUT_XLSX = OUT_DIR / "Panel_Descriptive_Statistics.xlsx"

df = pd.read_csv(PANEL)

# Numeric columns we care about (exclude IDs and categoricals)
NUM_COLS = [
    "Income_EUR",
    "Energy_Spend_Proxy_EUR",
    "Energy_Share_Pct",
    "Effective_Burden_Pct",
    "Inflation_YoY_Pct",
    "Elec_CPI_YoY",
    "Gas_CPI_YoY",
    "Energy_CPI_Composite_YoY",
    "Unemployment_Rate_Pct",
]

# Prettier display labels
VAR_LABELS = {
    "Income_EUR":                 "Income (€)",
    "Energy_Spend_Proxy_EUR":     "Energy Spend Proxy (€)",
    "Energy_Share_Pct":           "Energy Share (% of income)",
    "Effective_Burden_Pct":       "Effective Energy Burden (%)",
    "Inflation_YoY_Pct":          "Decile Inflation YoY (%)",
    "Elec_CPI_YoY":               "Electricity CPI YoY (%)",
    "Gas_CPI_YoY":                "Gas CPI YoY (%)",
    "Energy_CPI_Composite_YoY":   "Energy CPI Composite YoY (%)",
    "Unemployment_Rate_Pct":      "Unemployment Rate (%)",
}

# Observed-only subset (2017–2023, actual CSO EIHC01 data)
obs = df[df["Data_Source"] == "observed"].copy()

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
NAVY   = "1F4E79"
TEAL   = "2E75B6"
RED    = "C00000"
AMBER  = "F4B942"
GREEN  = "70AD47"
LGREY  = "F2F2F2"
DYEL   = "FFF2CC"

STATE_FILLS = {
    "Energy Poor": RED,
    "At Risk":     AMBER,
    "Not Poor":    GREEN,
}

thin  = Side(style="thin",   color="D9D9D9")
thick = Side(style="medium", color="AAAAAA")
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR   = Alignment(horizontal="center", vertical="center", wrap_text=False)
LFT   = Alignment(horizontal="left",   vertical="center")
RGT   = Alignment(horizontal="right",  vertical="center")


def hdr(ws, row, col, text, fg=NAVY, bold=True, wrap=True, align="center"):
    c = ws.cell(row, col, text)
    c.font      = Font(bold=bold, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=fg)
    c.border    = BDR
    c.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    return c


def val(ws, row, col, v, fmt=None, bold=False, bg=None, align="right"):
    cell = ws.cell(row, col, v)
    cell.font      = Font(bold=bold, size=10)
    cell.border    = BDR
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    return cell


def row_label(ws, row, col, text, bg=LGREY, bold=False):
    c = ws.cell(row, col, text)
    c.font      = Font(bold=bold, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.border    = BDR
    c.alignment = LFT
    return c


def set_col_width(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def section_hdr(ws, row, ncols, text, fg=TEAL):
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=ncols
    )
    c = ws.cell(row, 1, text)
    c.font      = Font(bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=fg)
    c.alignment = LFT
    c.border    = BDR


# ── STAT HELPER ───────────────────────────────────────────────────────────────
def pool_stats(series: pd.Series) -> dict:
    """Return a dict of summary stats for one numeric series."""
    s = series.dropna()
    if len(s) == 0:
        return {k: np.nan for k in
                ["N", "Mean", "SD", "Min", "P25", "Median", "P75", "Max", "Skew"]}
    return {
        "N":      len(s),
        "Mean":   round(s.mean(), 2),
        "SD":     round(s.std(ddof=1), 2),
        "Min":    round(s.min(), 2),
        "P25":    round(s.quantile(0.25), 2),
        "Median": round(s.median(), 2),
        "P75":    round(s.quantile(0.75), 2),
        "Max":    round(s.max(), 2),
        "Skew":   round(float(scipy_stats.skew(s, bias=False)), 3),
    }


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 1: Variable Summary (pooled)
# ═════════════════════════════════════════════════════════════════════════════
def build_table1(ws):
    """Full panel + observed-only pooled summary."""
    ws.title = "Table 1: Variable Summary"
    STAT_HDRS = ["N", "Mean", "SD", "Min", "P25", "Median", "P75", "Max", "Skew"]
    NCOLS     = 1 + len(STAT_HDRS)   # var label + stats

    ws.row_dimensions[1].height = 18

    # ── Section A: Full panel (all years, observed + interpolated) ──────────
    section_hdr(ws, 1, NCOLS, "Panel A: Full Panel  (2007–2023, N=170 decile-year obs.)")
    for ci, h in enumerate(["Variable"] + STAT_HDRS, start=1):
        hdr(ws, 2, ci, h)
    for ri, col in enumerate(NUM_COLS, start=3):
        s  = pool_stats(df[col])
        row_label(ws, ri, 1, VAR_LABELS[col])
        for ci, key in enumerate(STAT_HDRS, start=2):
            val(ws, ri, ci, s[key])

    gap = 3 + len(NUM_COLS) + 1   # blank row between sections

    # ── Section B: Observed only (2017–2023, actual decile inflation) ───────
    section_hdr(ws, gap, NCOLS,
                "Panel B: Observed Only  (2017–2023, N=70 decile-year obs.)",
                fg=TEAL)
    for ci, h in enumerate(["Variable"] + STAT_HDRS, start=1):
        hdr(ws, gap + 1, ci, h)
    for ri_off, col in enumerate(NUM_COLS, start=2):
        s  = pool_stats(obs[col])
        row_label(ws, gap + ri_off, 1, VAR_LABELS[col])
        for ci, key in enumerate(STAT_HDRS, start=2):
            val(ws, gap + ri_off, ci, s[key])

    # ── Note row ─────────────────────────────────────────────────────────────
    note_row = gap + 1 + len(NUM_COLS) + 1
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=NCOLS)
    nc = ws.cell(note_row, 1,
        "Notes: Panel A includes interpolated rows (2007–2016) where decile inflation is estimated "
        "by scaling aggregate housing CPI. Panel B restricted to observed CSO EIHC01 decile data "
        "(2017–2023). Skewness is bias-corrected (Fisher). Income extrapolated 2020–2023 via linear "
        "trend from 2015–2019 SILC data.")
    nc.font      = Font(italic=True, size=9)
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[note_row].height = 48

    set_col_width(ws, {
        "A": 30, "B": 6, "C": 10, "D": 10,
        "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 11,
    })
    ws.freeze_panes = "B3"


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 2: By State
# ═════════════════════════════════════════════════════════════════════════════
def build_table2(ws):
    ws.title = "Table 2: By State"
    STATES = ["Energy Poor", "At Risk", "Not Poor"]
    # Columns: Variable | Energy Poor (mean±SD, N) | At Risk (mean±SD, N) | Not Poor (mean±SD, N)
    # Laid out as: Variable | EP_N | EP_Mean | EP_SD | AR_N | AR_Mean | AR_SD | NP_N | NP_Mean | NP_SD
    STATE_COLS = ["N", "Mean", "Std Dev", "Median", "Min", "Max"]
    NCOLS = 1 + len(STATES) * len(STATE_COLS)

    # Merge header per state
    section_hdr(ws, 1, NCOLS,
                "Observed Panel (2017–2023): Key Variables by Energy Poverty State")
    col_ptr = 2
    for state in STATES:
        end_col = col_ptr + len(STATE_COLS) - 1
        ws.merge_cells(start_row=2, start_column=col_ptr,
                       end_row=2, end_column=end_col)
        fg = STATE_FILLS[state]
        c  = ws.cell(2, col_ptr, state)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=fg)
        c.alignment = CTR
        c.border    = BDR
        col_ptr     = end_col + 1

    # Sub-headers
    ws.cell(3, 1, "Variable").font = Font(bold=True, size=10)
    ws.cell(3, 1).fill      = PatternFill("solid", fgColor=NAVY)
    ws.cell(3, 1).alignment = LFT
    ws.cell(3, 1).border    = BDR
    ws.cell(3, 1).font      = Font(bold=True, color="FFFFFF", size=10)
    col_ptr = 2
    for state in STATES:
        fg = STATE_FILLS[state]
        for h in STATE_COLS:
            c = ws.cell(3, col_ptr, h)
            c.font      = Font(bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor=fg)
            c.alignment = CTR
            c.border    = BDR
            col_ptr    += 1

    # Data rows
    for ri, col in enumerate(NUM_COLS, start=4):
        row_label(ws, ri, 1, VAR_LABELS[col], bg=LGREY if ri % 2 == 0 else "FFFFFF")
        col_ptr = 2
        for state in STATES:
            grp  = obs[obs["State"] == state][col].dropna()
            vals = [
                len(grp),
                round(grp.mean(), 2)   if len(grp) else "",
                round(grp.std(ddof=1), 2) if len(grp) > 1 else "",
                round(grp.median(), 2) if len(grp) else "",
                round(grp.min(), 2)    if len(grp) else "",
                round(grp.max(), 2)    if len(grp) else "",
            ]
            bg = "FFF0F0" if state == "Energy Poor" else (
                 "FFFBE6" if state == "At Risk" else "F0FFF0")
            for v in vals:
                val(ws, ri, col_ptr, v, bg=bg if v != "" else "FFFFFF")
                col_ptr += 1

    # ── F-stat / ANOVA note ──────────────────────────────────────────────────
    note_row = 4 + len(NUM_COLS)
    # Run one-way ANOVA for each variable and print F, p
    anova_rows = [["Variable", "F-statistic", "p-value", "Interpretation"]]
    for col in NUM_COLS:
        groups = [obs[obs["State"] == st][col].dropna() for st in STATES]
        groups = [g for g in groups if len(g) >= 2]
        if len(groups) < 2:
            continue
        f, p = scipy_stats.f_oneway(*groups)
        interp = ("***" if p < 0.001 else "**" if p < 0.01
                  else "*" if p < 0.05 else "ns")
        anova_rows.append([VAR_LABELS[col], round(f, 3), round(p, 4), interp])

    section_hdr(ws, note_row + 1, 4,
                "ANOVA: test of mean differences across states (observed rows)", fg="595959")
    for ari, arow in enumerate(anova_rows, start=note_row + 2):
        for ci, av in enumerate(arow, start=1):
            cell = ws.cell(ari, ci, av)
            cell.font      = Font(bold=(ari == note_row + 2), size=10)
            cell.border    = BDR
            cell.alignment = CTR if ci > 1 else LFT
            if ari == note_row + 2:
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
            elif av in ("***", "**", "*"):
                cell.fill = PatternFill("solid", fgColor="FFE0E0")
            elif av == "ns":
                cell.fill = PatternFill("solid", fgColor=LGREY)

    set_col_width(ws, {
        "A": 30,
        **{get_column_letter(c): 10 for c in range(2, NCOLS + 2)},
    })
    ws.freeze_panes = "B4"


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 3: By Decile
# ═════════════════════════════════════════════════════════════════════════════
def build_table3(ws):
    ws.title = "Table 3: By Decile"
    SHOW_COLS = [
        "Income_EUR", "Energy_Share_Pct", "Effective_Burden_Pct",
        "Inflation_YoY_Pct", "Energy_CPI_Composite_YoY",
    ]
    STATS = ["Mean", "SD", "Min", "Max"]
    NCOLS = 1 + len(SHOW_COLS) * len(STATS)

    section_hdr(ws, 1, NCOLS,
                "Observed Panel (2017–2023): Key Variables by Income Decile (annual averages)")

    # Variable merge headers
    col_ptr = 2
    for col in SHOW_COLS:
        end = col_ptr + len(STATS) - 1
        ws.merge_cells(start_row=2, start_column=col_ptr,
                       end_row=2, end_column=end)
        c = ws.cell(2, col_ptr, VAR_LABELS[col])
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor=TEAL)
        c.alignment = CTR
        c.border    = BDR
        col_ptr     = end + 1

    # Sub-headers
    ws.cell(3, 1, "Decile").font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(3, 1).fill      = PatternFill("solid", fgColor=NAVY)
    ws.cell(3, 1).alignment = CTR
    ws.cell(3, 1).border    = BDR
    col_ptr = 2
    for col in SHOW_COLS:
        for h in STATS:
            c = ws.cell(3, col_ptr, h)
            c.font      = Font(bold=True, color="FFFFFF", size=9)
            c.fill      = PatternFill("solid", fgColor=TEAL)
            c.alignment = CTR
            c.border    = BDR
            col_ptr    += 1

    # Data
    for ri, d in enumerate(range(1, 11), start=4):
        bg = LGREY if ri % 2 == 0 else "FFFFFF"
        row_label(ws, ri, 1, f"Decile {d}", bg=bg, bold=(d <= 3))
        grp     = obs[obs["Decile"] == d]
        col_ptr = 2
        for col in SHOW_COLS:
            s    = grp[col].dropna()
            vals = [
                round(s.mean(), 2)       if len(s) else "",
                round(s.std(ddof=1), 2)  if len(s) > 1 else "",
                round(s.min(), 2)        if len(s) else "",
                round(s.max(), 2)        if len(s) else "",
            ]
            for v in vals:
                cell = val(ws, ri, col_ptr, v, bg=bg)
                col_ptr += 1

    set_col_width(ws, {"A": 12,
                       **{get_column_letter(c): 9 for c in range(2, NCOLS + 2)}})
    ws.freeze_panes = "B4"


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 4: Annual Trends
# ═════════════════════════════════════════════════════════════════════════════
def build_table4(ws):
    ws.title = "Table 4: Annual Trends"
    HDR_COLS = [
        "Year",
        "Mean Income (€)",
        "Mean Energy Burden (%)",
        "Mean Inflation YoY (%)",
        "Mean Energy CPI (%)",
        "% Energy Poor",
        "% At Risk",
        "% Not Poor",
        "N Deciles",
    ]
    NCOLS = len(HDR_COLS)

    section_hdr(ws, 1, NCOLS,
                "Annual Panel Aggregates: All Deciles (observed 2017–2023 only)")
    for ci, h in enumerate(HDR_COLS, start=1):
        hdr(ws, 2, ci, h)

    years = sorted(obs["Year"].unique())
    for ri, yr in enumerate(years, start=3):
        ydf  = obs[obs["Year"] == yr]
        n    = len(ydf)
        bg   = LGREY if ri % 2 == 0 else "FFFFFF"

        pct = {st: round(100 * (ydf["State"] == st).sum() / n, 1) for st in
               ["Energy Poor", "At Risk", "Not Poor"]}

        row_vals = [
            yr,
            round(ydf["Income_EUR"].mean(), 0),
            round(ydf["Effective_Burden_Pct"].mean(), 2),
            round(ydf["Inflation_YoY_Pct"].mean(), 2),
            round(ydf["Energy_CPI_Composite_YoY"].mean(), 2),
            pct["Energy Poor"],
            pct["At Risk"],
            pct["Not Poor"],
            n,
        ]
        for ci, v in enumerate(row_vals, start=1):
            cell = val(ws, ri, ci, v, bg=bg,
                       align="center" if ci == 1 else "right")
            # Highlight energy poor % cells
            if ci == 6 and isinstance(v, float) and v > 0:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
            if ci == 7 and isinstance(v, float) and v > 0:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")

    # Totals row
    tr = 3 + len(years)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=1)
    c = ws.cell(tr, 1, "All years (pooled)")
    c.font = Font(bold=True, size=10); c.border = BDR; c.alignment = LFT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.font = Font(bold=True, color="FFFFFF", size=10)

    pool_vals = [
        round(obs["Income_EUR"].mean(), 0),
        round(obs["Effective_Burden_Pct"].mean(), 2),
        round(obs["Inflation_YoY_Pct"].mean(), 2),
        round(obs["Energy_CPI_Composite_YoY"].mean(), 2),
        round(100 * (obs["State"] == "Energy Poor").mean(), 1),
        round(100 * (obs["State"] == "At Risk").mean(), 1),
        round(100 * (obs["State"] == "Not Poor").mean(), 1),
        len(obs),
    ]
    for ci, v in enumerate(pool_vals, start=2):
        c = val(ws, tr, ci, v, bold=True, bg=NAVY)
        c.font = Font(bold=True, color="FFFFFF", size=10)

    set_col_width(ws, {
        "A": 8, "B": 18, "C": 22, "D": 22, "E": 22,
        "F": 14, "G": 12, "H": 12, "I": 12,
    })
    ws.freeze_panes = "A3"


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 5: State Frequency (transition matrix raw material)
# ═════════════════════════════════════════════════════════════════════════════
def build_table5(ws):
    ws.title = "Table 5: State Frequency"
    STATES = ["Energy Poor", "At Risk", "Not Poor"]

    # Part A: counts per year × state
    section_hdr(ws, 1, 5,
                "Part A: Decile-Year Count by State and Year (observed 2017–2023)")
    for ci, h in enumerate(["Year", "Energy Poor", "At Risk", "Not Poor", "Total"], start=1):
        hdr(ws, 2, ci, h)

    years = sorted(obs["Year"].unique())
    for ri, yr in enumerate(years, start=3):
        ydf = obs[obs["Year"] == yr]
        bg  = LGREY if ri % 2 == 0 else "FFFFFF"
        val(ws, ri, 1, yr,   bg=bg, align="center")
        for ci, st in enumerate(STATES, start=2):
            n = (ydf["State"] == st).sum()
            cell = val(ws, ri, ci, int(n), bg=bg)
            if st == "Energy Poor" and n > 0:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
            if st == "At Risk" and n > 0:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
        val(ws, ri, 5, len(ydf), bg=bg, bold=True)

    # Totals
    tr = 3 + len(years)
    val(ws, tr, 1, "Total", bold=True, bg=NAVY, align="center")
    ws.cell(tr, 1).font = Font(bold=True, color="FFFFFF", size=10)
    for ci, st in enumerate(STATES, start=2):
        n = (obs["State"] == st).sum()
        c = val(ws, tr, ci, int(n), bold=True, bg=NAVY)
        c.font = Font(bold=True, color="FFFFFF", size=10)
    c = val(ws, tr, 5, len(obs), bold=True, bg=NAVY)
    c.font = Font(bold=True, color="FFFFFF", size=10)

    # Part B: same as % of row total
    gap = tr + 2
    section_hdr(ws, gap, 5,
                "Part B: Row % (share of deciles in each state, per year)")
    for ci, h in enumerate(["Year", "% Energy Poor", "% At Risk", "% Not Poor", "Total %"], start=1):
        hdr(ws, gap + 1, ci, h)

    for ri, yr in enumerate(years, start=gap + 2):
        ydf = obs[obs["Year"] == yr]
        n   = len(ydf)
        bg  = LGREY if ri % 2 == 0 else "FFFFFF"
        val(ws, ri, 1, yr, bg=bg, align="center")
        for ci, st in enumerate(STATES, start=2):
            pct = round(100 * (ydf["State"] == st).sum() / n, 1)
            cell = val(ws, ri, ci, pct, bg=bg)
            if st == "Energy Poor" and pct > 0:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
            if st == "At Risk" and pct > 0:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
        val(ws, ri, 5, 100.0, bg=bg, bold=True)

    # Part C: by decile
    gap2 = gap + 2 + len(years) + 2
    section_hdr(ws, gap2, 5,
                "Part C: Decile-Year Count by State and Decile (observed 2017–2023)")
    for ci, h in enumerate(["Decile", "Energy Poor", "At Risk", "Not Poor", "Total"], start=1):
        hdr(ws, gap2 + 1, ci, h)

    for ri, d in enumerate(range(1, 11), start=gap2 + 2):
        ddf = obs[obs["Decile"] == d]
        bg  = LGREY if ri % 2 == 0 else "FFFFFF"
        val(ws, ri, 1, f"Decile {d}", bg=bg, align="center")
        for ci, st in enumerate(STATES, start=2):
            n = (ddf["State"] == st).sum()
            cell = val(ws, ri, ci, int(n), bg=bg)
            if st == "Energy Poor" and n > 0:
                cell.fill = PatternFill("solid", fgColor="FFD7D7")
            if st == "At Risk" and n > 0:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
        val(ws, ri, 5, len(ddf), bg=bg, bold=True)

    set_col_width(ws, {"A": 14, "B": 16, "C": 12, "D": 12, "E": 10})
    ws.freeze_panes = "A3"


# ═════════════════════════════════════════════════════════════════════════════
#  TABLE 6: Correlation Matrix
# ═════════════════════════════════════════════════════════════════════════════
def build_table6(ws):
    ws.title = "Table 6: Correlation Matrix"
    CORR_COLS = [c for c in NUM_COLS if obs[c].notna().sum() > 10]
    labels    = [VAR_LABELS[c] for c in CORR_COLS]
    n_vars    = len(CORR_COLS)
    NCOLS     = 1 + n_vars

    section_hdr(ws, 1, NCOLS,
                "Pearson Correlation Matrix: Observed Panel (2017–2023, pairwise complete obs.)")

    # Column headers
    ws.cell(2, 1, "Variable").fill      = PatternFill("solid", fgColor=NAVY)
    ws.cell(2, 1).font      = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(2, 1).border    = BDR
    ws.cell(2, 1).alignment = LFT
    for ci, lbl in enumerate(labels, start=2):
        hdr(ws, 2, ci, lbl, wrap=True)
        ws.row_dimensions[2].height = 48

    corr_data = obs[CORR_COLS].copy()

    for ri, (row_col, row_lbl) in enumerate(zip(CORR_COLS, labels), start=3):
        bg = LGREY if ri % 2 == 0 else "FFFFFF"
        row_label(ws, ri, 1, row_lbl, bg=bg)
        for ci, col_col in enumerate(CORR_COLS, start=2):
            if row_col == col_col:
                cell = val(ws, ri, ci, 1.000, bg="D9D9D9")
                cell.font = Font(bold=True, size=10)
                continue
            pair = corr_data[[row_col, col_col]].dropna()
            if len(pair) < 3:
                val(ws, ri, ci, ":", bg=bg, align="center")
                continue
            result = scipy_stats.pearsonr(pair[row_col], pair[col_col])
            r      = round(float(np.float64(result.correlation)), 3)
            p      = float(np.float64(result.pvalue))
            cell = val(ws, ri, ci, r, bg=bg)
            # Colour coding: strong positive = blue, strong negative = red
            if abs(r) >= 0.7:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=("DBEAFE" if r > 0 else "FEE2E2")
                )
                cell.font = Font(bold=True, size=10)
            # Add significance stars as comment-style suffix
            star = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""
            if star:
                cell.value = f"{r}{star}"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    note_row = 3 + n_vars + 1
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=NCOLS)
    nc = ws.cell(note_row, 1,
        "Notes: * p<0.05  ** p<0.01  *** p<0.001 (two-tailed). "
        "Blue shading = strong positive correlation (|r| ≥ 0.7). "
        "Red shading = strong negative correlation. "
        "Diagonal = 1.000 by definition. "
        "Correlations based on observed decile-years only (2017–2023, N≤70).")
    nc.font      = Font(italic=True, size=9)
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[note_row].height = 36

    set_col_width(ws, {
        "A": 30,
        **{get_column_letter(c): 12 for c in range(2, NCOLS + 2)},
    })
    ws.freeze_panes = "B3"


# ═════════════════════════════════════════════════════════════════════════════
#  ASSEMBLE WORKBOOK
# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# Build all sheets
ws1 = wb.active
build_table1(ws1)

ws2 = wb.create_sheet()
build_table2(ws2)

ws3 = wb.create_sheet()
build_table3(ws3)

ws4 = wb.create_sheet()
build_table4(ws4)

ws5 = wb.create_sheet()
build_table5(ws5)

ws6 = wb.create_sheet()
build_table6(ws6)

wb.save(OUT_XLSX)

# ── CONSOLE SUMMARY ───────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("  Panel Descriptive Statistics: Export Complete")
print(f"{'='*60}")
print(f"  Output : {OUT_XLSX}")
print(f"  Sheets : 6")
print(f"\n  Table 1  Variable Summary   (pooled + observed-only)")
print(f"  Table 2  By State           (Energy Poor / At Risk / Not Poor)")
print(f"  Table 3  By Decile          (D1–D10 panel averages)")
print(f"  Table 4  Annual Trends      (year-level aggregates)")
print(f"  Table 5  State Frequency    (counts & % per year & decile)")
print(f"  Table 6  Correlation Matrix (Pearson, significance-starred)")
print(f"\n  Panel:  {len(df)} obs total  |  {len(obs)} observed  |  {len(df)-len(obs)} interpolated")
print(f"  States: Energy Poor={( df['State']=='Energy Poor').sum()}  "
      f"At Risk={(df['State']=='At Risk').sum()}  "
      f"Not Poor={(df['State']=='Not Poor').sum()}")
print(f"\n✅  Done")
