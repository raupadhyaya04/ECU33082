"""
Export Descriptive Statistics — Document-Ready Format
======================================================
Produces TWO narrow sheets designed to be copy-pasted into a document:

  Sheet 1 "Summary Stats"   (12 cols) — the write-up table:
      Variable | N | Mean | Trim.Mean | Median | Std Dev | CV% | IQR | MAD
               | Skewness | Kurtosis | Shape note

  Sheet 2 "Percentile Table" (11 cols) — the position/frequency table:
      Variable | N | N+ | N- | %+ | %-
               | Min | P10 | P25 | Median | P75 | P90 | Max

Both sheets use the same section headings and row order.
Rounding: 2 d.p. for rates/%, 0 d.p. for € values, 4 d.p. for prices.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path("Cleaned Data")
OUT_PATH = DATA_DIR / "Descriptive_Statistics_Document.xlsx"

# ── stat engine ───────────────────────────────────────────────────────────────

def full_stats(s: pd.Series) -> dict:
    s = s.dropna().reset_index(drop=True)
    n = len(s)
    if n == 0:
        return {}
    p10, p25, p75, p90 = s.quantile([0.10, 0.25, 0.75, 0.90]).values
    iqr     = p75 - p25
    mad     = float(np.median(np.abs(s - s.median())))
    trimmed = float(scipy_stats.trim_mean(s.values, 0.05))
    cv      = float(s.std() / s.mean() * 100) if s.mean() != 0 else np.nan
    n_pos   = int((s > 0).sum())
    n_neg   = int((s < 0).sum())
    return {
        "N":         n,
        "N+":        n_pos,
        "N-":        n_neg,
        "%+":        n_pos / n * 100,
        "%-":        n_neg / n * 100,
        "Mean":      float(s.mean()),
        "Trim.Mean": trimmed,
        "Median":    float(s.median()),
        "Std Dev":   float(s.std()),
        "CV%":       cv,
        "IQR":       float(iqr),
        "MAD":       mad,
        "Skewness":  float(s.skew()),
        "Kurtosis":  float(s.kurt()),
        "Min":       float(s.min()),
        "P10":       float(p10),
        "P25":       float(p25),
        "P75":       float(p75),
        "P90":       float(p90),
        "Max":       float(s.max()),
    }

# Sheet 1 columns — central tendency + dispersion + shape  (13 cols total)
S1_COLS = ["N", "Mean", "Trim.Mean", "Median", "Std Dev", "CV%",
           "IQR", "MAD", "Skewness", "Kurtosis"]

# Sheet 2 columns — frequency + position  (10 cols total)
S2_COLS = ["N", "N+", "N-", "%+", "%-", "Min", "P10", "P25", "P75", "P90", "Max"]

def r(val, rule):
    """Round a value according to the format rule."""
    if val == "" or (isinstance(val, float) and np.isnan(val)):
        return ""
    if rule == "int":  return int(round(val))
    if rule == "price": return round(val, 4)
    if rule == "eur":  return int(round(val))
    return round(val, 2)   # rate / pct default

def make_row(stats: dict, cols: list, rule: str) -> list:
    out = []
    for col in cols:
        v = stats.get(col, "")
        if v == "":
            out.append("")
        elif col in ("N", "N+", "N-"):
            out.append(int(v))
        elif col in ("%+", "%-"):
            out.append(round(v, 1))
        else:
            out.append(r(v, rule))
    return out

# ── load data ─────────────────────────────────────────────────────────────────

housing_cpi = pd.read_csv(DATA_DIR / "CPI_Housing_2007_2023.csv")
cpi_broad   = pd.read_csv(DATA_DIR / "CPI_AllItems_and_Housing_Monthly_2007_2023.csv")
cpi_broad["Month_dt"] = pd.to_datetime(cpi_broad["Month_dt"])
decile_inf  = pd.read_csv(DATA_DIR / "Inflation_by_Income_Decile_2017_2023.csv")
decile_inf["Month_dt"] = pd.to_datetime(decile_inf["Month_dt"])
energy_cpi  = pd.read_csv(DATA_DIR / "CPI_Energy_Products_2007_2023.csv")
gas         = pd.read_csv(DATA_DIR / "Gas_Prices_Residential_2015_2023.csv")
elec        = pd.read_csv(DATA_DIR / "Electricity_Prices_Residential_2015_2023.csv")
income      = pd.read_csv(DATA_DIR / "Income_Median_by_AgeGroup_2007_2023.csv")
income_vol  = pd.read_csv(DATA_DIR / "Income_Volatility_YoY_by_AgeGroup_2007_2023.csv")
tenure      = pd.read_csv(DATA_DIR / "Disposable_Income_by_Tenure_2007_2023.csv")
emp_q       = pd.read_csv(DATA_DIR / "Employment_Rates_Quarterly_ILO_2007_2017.csv")
unemp       = pd.read_csv(DATA_DIR / "Unemployment_Rate_Monthly_2007_2023.csv")
poverty     = pd.read_csv(DATA_DIR / "Poverty_Rates_2007_2023.csv")

energy_wide = energy_cpi.pivot(index="Year", columns="Energy_Product", values="CPI_Index")
energy_yoy  = (energy_wide.pct_change() * 100).reset_index().melt(
    id_vars="Year", var_name="Energy_Product", value_name="YoY_Pct_Change").dropna()

no_all = decile_inf[decile_inf["Income_Decile"] != "All deciles"]
spread = no_all.groupby("Month_dt")["YoY_Inflation_Pct"].agg(
    Spread=lambda x: x.max() - x.min()).reset_index()

vol_long = income_vol.melt(id_vars="Year", var_name="Age_Group",
                           value_name="YoY_Pct_Change").dropna()

# ── variable sections ─────────────────────────────────────────────────────────
# (label, series, fmt_rule)

SECTIONS = [
    {
        "heading": "1.  INFLATION — Housing & All-Items CPI",
        "note":    "Monthly, 2007–2023  |  Source: CSO CPM01, CPM13",
        "rows": [
            ("Housing CPI – YoY % Change",
             housing_cpi["YoY_Pct_Change"], "rate"),
            ("Housing CPI – Index Level (Base Dec 2023=100)",
             housing_cpi["CPI_Index"], "rate"),
            ("All-Items CPI – YoY % Change",
             cpi_broad.loc[cpi_broad["Commodity_Group"] == "All items",
                           "YoY_Inflation_Pct"], "rate"),
            ("Housing CPI YoY % (CPM13 series)",
             cpi_broad.loc[cpi_broad["Commodity_Group"] != "All items",
                           "YoY_Inflation_Pct"], "rate"),
        ],
    },
    {
        "heading": "2.  INFLATION BY INCOME DECILE",
        "note":    "Monthly, 2017–2023  |  Source: CSO EIHC01  |  N=70 per decile",
        "rows": [
            ("1st Decile (Lowest Income) – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "1st decile", "YoY_Inflation_Pct"], "rate"),
            ("2nd Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "2nd decile", "YoY_Inflation_Pct"], "rate"),
            ("3rd Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "3rd decile", "YoY_Inflation_Pct"], "rate"),
            ("4th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "4th decile", "YoY_Inflation_Pct"], "rate"),
            ("5th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "5th decile", "YoY_Inflation_Pct"], "rate"),
            ("6th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "6th decile", "YoY_Inflation_Pct"], "rate"),
            ("7th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "7th decile", "YoY_Inflation_Pct"], "rate"),
            ("8th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "8th decile", "YoY_Inflation_Pct"], "rate"),
            ("9th Decile – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "9th decile", "YoY_Inflation_Pct"], "rate"),
            ("10th Decile (Highest Income) – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "10th decile", "YoY_Inflation_Pct"], "rate"),
            ("All Deciles – YoY Inflation %",
             decile_inf.loc[decile_inf["Income_Decile"] == "All deciles", "YoY_Inflation_Pct"], "rate"),
            ("Cross-Decile Spread (Max − Min, pp)",
             spread["Spread"], "rate"),
        ],
    },
    {
        "heading": "3.  ENERGY CPI — Index Level & YoY % Change by Product",
        "note":    "Annual, 2007–2023  |  Source: CSO EIIEEA04  |  N=17 (index), N=16 (YoY)",
        "rows": [
            ("Electricity – CPI Index Level",
             energy_cpi.loc[energy_cpi["Energy_Product"] == "Electricity", "CPI_Index"], "rate"),
            ("Electricity – YoY % Change",
             energy_yoy.loc[energy_yoy["Energy_Product"] == "Electricity", "YoY_Pct_Change"], "rate"),
            ("Natural Gas – CPI Index Level",
             energy_cpi.loc[energy_cpi["Energy_Product"] == "Natural gas", "CPI_Index"], "rate"),
            ("Natural Gas – YoY % Change",
             energy_yoy.loc[energy_yoy["Energy_Product"] == "Natural gas", "YoY_Pct_Change"], "rate"),
            ("Solid Fuel (coal/peat/wood) – CPI Index Level",
             energy_cpi.loc[energy_cpi["Energy_Product"].str.startswith("Solid"), "CPI_Index"], "rate"),
            ("Solid Fuel – YoY % Change",
             energy_yoy.loc[energy_yoy["Energy_Product"].str.startswith("Solid"), "YoY_Pct_Change"], "rate"),
            ("Autodiesel – CPI Index Level",
             energy_cpi.loc[energy_cpi["Energy_Product"] == "Autodiesel", "CPI_Index"], "rate"),
            ("Autodiesel – YoY % Change",
             energy_yoy.loc[energy_yoy["Energy_Product"] == "Autodiesel", "YoY_Pct_Change"], "rate"),
            ("Petrol – CPI Index Level",
             energy_cpi.loc[energy_cpi["Energy_Product"] == "Petrol", "CPI_Index"], "rate"),
            ("Petrol – YoY % Change",
             energy_yoy.loc[energy_yoy["Energy_Product"] == "Petrol", "YoY_Pct_Change"], "rate"),
        ],
    },
    {
        "heading": "4.  RESIDENTIAL ENERGY PRICES (Actual)",
        "note":    "Semi-annual, 2015–2023  |  Source: CSO TMEGB01/03  |  N=18  |  All-in incl. taxes",
        "rows": [
            ("Residential Gas Price (€/GJ, band D2)",
             gas["Price_EUR_per_GJ"], "price"),
            ("Residential Electricity Price (€/kWh, band DC)",
             elec["Price_EUR_per_kWh"], "price"),
        ],
    },
    {
        "heading": "5.  INCOME — Median Real Disposable Income",
        "note":    "Annual, 2007–2023  |  Source: CSO SIA13, SIA52  |  Constant euros  |  N=13",
        "rows": [
            ("18–64 Years – Median Real Disposable Income (€)",
             income.loc[income["Age_Group"] == "18 - 64 years",
                        "Median_Real_Disposable_Income_EUR"], "eur"),
            ("65+ Years – Median Real Disposable Income (€)",
             income.loc[income["Age_Group"] == "65 years and over",
                        "Median_Real_Disposable_Income_EUR"], "eur"),
            ("18–64 Years – Income YoY % Change",
             vol_long.loc[vol_long["Age_Group"].str.contains("18"), "YoY_Pct_Change"], "rate"),
            ("65+ Years – Income YoY % Change",
             vol_long.loc[vol_long["Age_Group"].str.contains("65"), "YoY_Pct_Change"], "rate"),
            ("Owner-Occupied – Median Real Disposable Income (€)",
             tenure.loc[tenure["Tenure_Type"] == "Owner-occupied",
                        "Median_Real_Disposable_Income_EUR"], "eur"),
            ("Rented / Rent-Free – Median Real Disposable Income (€)",
             tenure.loc[tenure["Tenure_Type"] == "Rented or rent free",
                        "Median_Real_Disposable_Income_EUR"], "eur"),
        ],
    },
    {
        "heading": "6.  EMPLOYMENT & UNEMPLOYMENT",
        "note":    "Quarterly ILO 2007–2017 (CSO QNQ20, N=42);  Monthly SA 2007–2023 (CSO LRM03, N=100)",
        "rows": [
            ("ILO Participation Rate – 15+ (%)",
             emp_q.loc[emp_q["Statistic"] == "ILO Participation Rates (15 years and over)",
                       "Rate_Pct"], "rate"),
            ("ILO Participation Rate – 15+ Seasonally Adjusted (%)",
             emp_q.loc[emp_q["Statistic"] ==
                       "ILO Participation Rates (15 years and over) (Seasonally Adjusted)",
                       "Rate_Pct"], "rate"),
            ("ILO Unemployment Rate – 15–74 (%)",
             emp_q.loc[emp_q["Statistic"] == "ILO Unemployment Rates (15 - 74 years)",
                       "Rate_Pct"], "rate"),
            ("ILO Unemployment Rate – 15–74 Seasonally Adjusted (%)",
             emp_q.loc[emp_q["Statistic"] ==
                       "ILO Unemployment Rates (15 - 74 years) (Seasonally Adjusted)",
                       "Rate_Pct"], "rate"),
            ("Unemployment Rate – Monthly SA (%)",
             unemp["Unemployment_Rate_Pct"], "rate"),
        ],
    },
    {
        "heading": "7.  POVERTY RATES",
        "note":    "Annual (CSO SIA01)  |  ⚠ API currently returns 2007 only",
        "rows": [
            ("At-Risk-of-Poverty Rate – 60% Median Income (%)",
             poverty.loc[poverty["Statistic"].str.contains("At Risk", na=False), "VALUE"], "rate"),
            ("Consistent Poverty Rate – 11 New Indicators (%)",
             poverty.loc[poverty["Statistic"].str.contains("11 New", na=False), "VALUE"], "rate"),
            ("Consistent Poverty Rate – 8 Old Indicators (%)",
             poverty.loc[poverty["Statistic"].str.contains("1 of 8", na=False), "VALUE"], "rate"),
        ],
    },
]

# ── shared style helpers ──────────────────────────────────────────────────────

DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"

def filled(h): return PatternFill("solid", fgColor=h)
def bfont(colour=WHITE, sz=10): return Font(bold=True, color=colour, size=sz)
def nfont(sz=10): return Font(size=sz)

_thin = Side(style="thin", color="BFBFBF")
bdr   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def write_sheet(wb: Workbook, sheet_title: str, stat_cols: list,
                col_group_label: str, data_col_w: int = 10):
    """
    Write one sheet.  Returns the sheet object.
    stat_cols  : list of keys from full_stats() to include
    """
    ws = wb.create_sheet(sheet_title)
    NCOLS = 1 + len(stat_cols)

    # ── title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    tc = ws.cell(1, 1,
        value="Descriptive Statistics — Energy Poverty Transition Analysis"
              f"  [{col_group_label}]  (Ireland, 2007–2023)")
    tc.font = Font(bold=True, color=WHITE, size=12)
    tc.fill = filled(DARK_BLUE)
    tc.alignment = CTR
    ws.row_dimensions[1].height = 20

    # ── column header
    ws.cell(2, 1, "Variable / Series").font      = bfont(WHITE, 10)
    ws.cell(2, 1).fill      = filled(MID_BLUE)
    ws.cell(2, 1).alignment = LFT
    ws.cell(2, 1).border    = bdr

    for ci, col in enumerate(stat_cols, start=2):
        c = ws.cell(2, ci, col)
        c.font = bfont(WHITE, 10)
        c.fill = filled(MID_BLUE)
        c.alignment = CTR
        c.border = bdr
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "B3"

    # ── column widths
    ws.column_dimensions["A"].width = 46
    for i in range(2, NCOLS + 1):
        ws.column_dimensions[get_column_letter(i)].width = data_col_w

    current_row = 3

    for sec in SECTIONS:
        # section heading
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=NCOLS)
        hc = ws.cell(current_row, 1, sec["heading"])
        hc.font = Font(bold=True, color=WHITE, size=10)
        hc.fill = filled(DARK_BLUE)
        hc.alignment = LFT
        hc.border = bdr
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # note
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=NCOLS)
        nc = ws.cell(current_row, 1, sec["note"])
        nc.font = Font(italic=True, color="595959", size=9)
        nc.fill = filled(LIGHT_BLUE)
        nc.alignment = LFT
        nc.border = bdr
        ws.row_dimensions[current_row].height = 13
        current_row += 1

        # data rows
        for ri, (label, series, rule) in enumerate(sec["rows"]):
            bg   = WHITE if ri % 2 == 0 else LIGHT_GREY
            stats = full_stats(series)
            vals  = make_row(stats, stat_cols, rule)

            vc = ws.cell(current_row, 1, label)
            vc.font = nfont(); vc.fill = filled(bg)
            vc.alignment = LFT; vc.border = bdr

            for ci, val in enumerate(vals, start=2):
                sc = ws.cell(current_row, ci, val)
                sc.font = nfont(); sc.fill = filled(bg)
                sc.alignment = CTR; sc.border = bdr

            ws.row_dimensions[current_row].height = 14
            current_row += 1

        current_row += 1   # blank gap between sections

    return ws, current_row


# ── build workbook ────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

ws1, r1 = write_sheet(
    wb,
    sheet_title     = "Central Tendency & Dispersion",
    stat_cols       = S1_COLS,
    col_group_label = "Central Tendency · Dispersion · Shape",
    data_col_w      = 11,
)

ws2, r2 = write_sheet(
    wb,
    sheet_title     = "Frequency & Position",
    stat_cols       = S2_COLS,
    col_group_label = "Frequency · Position (Percentiles)",
    data_col_w      = 10,
)

wb.save(OUT_PATH)

n_vars = sum(len(s["rows"]) for s in SECTIONS)
print(f"\n✅  Exported → {OUT_PATH}")
print(f"   Sheet 1 'Central Tendency & Dispersion'  —  {len(S1_COLS)+1} columns  ({r1-1} rows)")
print(f"   Sheet 2 'Frequency & Position'           —  {len(S2_COLS)+1} columns  ({r2-1} rows)")
print(f"   {n_vars} variable rows across {len(SECTIONS)} sections")
