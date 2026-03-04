"""
Descriptive Statistics for Energy Poverty Transition Analysis
=============================================================
Project : Analysis of Energy Poverty Transitions with inflation as main comparator
Stats   : Central Tendency · Dispersion · Position · Frequency / Shape
"""

import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats as scipy_stats

DATA_DIR = Path("Cleaned Data")
OUT_DIR  = Path("Cleaned Data")

# ── helpers ───────────────────────────────────────────────────────────────────

def section(title: str):
    bar = "=" * 70
    print(f"\n{bar}\n  {title}\n{bar}")


def _full_stats(s: pd.Series, label: str) -> dict:
    """
    Compute a full suite of descriptive statistics for a numeric Series.

    Central Tendency  : Mean, Trimmed Mean (5%), Median, Mode
    Dispersion        : Std Dev, Variance, CV, Range, IQR, MAD
    Position          : P5, P10, P25, P75, P90, P95
    Frequency / Shape : Skewness, Kurtosis (excess), N_pos, N_neg, N_zero,
                        Pct_positive, Pct_negative
    """
    s = s.dropna()
    n = len(s)
    if n == 0:
        return {"Variable": label, "N": 0}

    q = [0.05, 0.10, 0.25, 0.75, 0.90, 0.95]
    p5, p10, p25, p75, p90, p95 = s.quantile(q).values

    iqr       = p75 - p25
    mad       = float(np.median(np.abs(s - s.median())))
    trimmed   = float(scipy_stats.trim_mean(s.values, 0.05))
    mode_vals = s.mode()
    mode_val  = float(mode_vals.iloc[0]) if len(mode_vals) > 0 else np.nan
    cv        = float(s.std() / s.mean() * 100) if s.mean() != 0 else np.nan
    skew      = float(s.skew())
    kurt      = float(s.kurt())          # excess kurtosis (normal = 0)

    n_pos  = int((s > 0).sum())
    n_neg  = int((s < 0).sum())
    n_zero = int((s == 0).sum())

    return {
        "Variable":       label,
        # ── Frequency ────────────────────────────────────────────────────────
        "N":              n,
        "N_Missing":      int(s.isna().sum()),   # already dropped, kept for reference
        "N_Positive":     n_pos,
        "N_Negative":     n_neg,
        "N_Zero":         n_zero,
        "Pct_Positive":   round(n_pos  / n * 100, 2),
        "Pct_Negative":   round(n_neg  / n * 100, 2),
        # ── Central Tendency ─────────────────────────────────────────────────
        "Mean":           round(float(s.mean()), 4),
        "Trimmed_Mean_5": round(trimmed, 4),
        "Median":         round(float(s.median()), 4),
        "Mode":           round(mode_val, 4),
        # ── Dispersion ───────────────────────────────────────────────────────
        "Std_Dev":        round(float(s.std()), 4),
        "Variance":       round(float(s.var()), 4),
        "CV_Pct":         round(cv, 2),
        "Range":          round(float(s.max() - s.min()), 4),
        "IQR":            round(float(iqr), 4),
        "MAD":            round(mad, 4),
        # ── Position ─────────────────────────────────────────────────────────
        "Min":            round(float(s.min()), 4),
        "P5":             round(float(p5), 4),
        "P10":            round(float(p10), 4),
        "P25":            round(float(p25), 4),
        "P75":            round(float(p75), 4),
        "P90":            round(float(p90), 4),
        "P95":            round(float(p95), 4),
        "Max":            round(float(s.max()), 4),
        # ── Shape ────────────────────────────────────────────────────────────
        "Skewness":       round(skew, 4),
        "Kurtosis_Excess":round(kurt, 4),
    }


def desc(df: pd.DataFrame, value_col: str,
         group_cols: list[str] | None = None,
         label: str = "") -> pd.DataFrame:
    """
    Full descriptive stats table, optionally grouped by one or more columns.
    When grouped, each group gets its own row of full stats.
    """
    if group_cols:
        rows = []
        for group_key, grp in df.groupby(group_cols):
            # Build a readable label from group key
            if isinstance(group_key, tuple):
                grp_label = " | ".join(str(k) for k in group_key)
            else:
                grp_label = str(group_key)
            row = _full_stats(grp[value_col], grp_label)
            # Prepend the group column values for clarity
            if isinstance(group_key, tuple):
                for col, val in zip(group_cols, group_key):
                    row[col] = val
            else:
                row[group_cols[0]] = group_key
            rows.append(row)

        # Reorder: group cols first, then stats
        stat_cols = [c for c in rows[0].keys() if c not in group_cols and c != "Variable"]
        col_order = group_cols + stat_cols
        result = pd.DataFrame(rows)[col_order]
    else:
        row = _full_stats(df[value_col], label or value_col)
        row.pop("Variable", None)
        result = pd.DataFrame([{"Variable": label or value_col, **row}])

    return result


all_stats: dict[str, pd.DataFrame] = {}

# ══════════════════════════════════════════════════════════════════════════════
# 1. INFLATION — Housing CPI (monthly, 2007–2023)
#    Primary comparator for energy poverty transitions
# ══════════════════════════════════════════════════════════════════════════════
section("1. INFLATION — Housing CPI (Monthly, 2007–2023)")
housing_cpi = pd.read_csv(DATA_DIR / "CPI_Housing_2007_2023.csv")
housing_cpi["Month_dt"] = pd.to_datetime(housing_cpi["Month"], format="%Y %B")

# Overall
d1a = desc(housing_cpi, "YoY_Pct_Change", label="Housing CPI – YoY % Change")
d1b = desc(housing_cpi, "CPI_Index",      label="Housing CPI – Index Level")
d1_overall = pd.concat([d1a, d1b], ignore_index=True)
print("\n[Overall]")
print(d1_overall.to_string(index=False))

# By year (annual average of monthly readings)
d1_annual = (
    housing_cpi.groupby("Year")[["YoY_Pct_Change", "CPI_Index"]]
    .mean()
    .round(3)
    .reset_index()
    .rename(columns={"YoY_Pct_Change": "Avg_YoY_Pct_Change",
                     "CPI_Index":      "Avg_CPI_Index"})
)
print("\n[Annual Averages]")
print(d1_annual.to_string(index=False))
all_stats["1_Housing_CPI_Overall"]       = d1_overall
all_stats["1_Housing_CPI_Annual_Avg"]    = d1_annual

# ══════════════════════════════════════════════════════════════════════════════
# 2. INFLATION — All Items & Housing CPI (monthly, 2007–2023)
# ══════════════════════════════════════════════════════════════════════════════
section("2. INFLATION — All Items vs Housing CPI YoY % (Monthly, 2007–2023)")
cpi_broad = pd.read_csv(DATA_DIR / "CPI_AllItems_and_Housing_Monthly_2007_2023.csv")
cpi_broad["Month_dt"] = pd.to_datetime(cpi_broad["Month_dt"])

d2 = desc(cpi_broad, "YoY_Inflation_Pct",
          group_cols=["Commodity_Group"],
          label="YoY Inflation %")
print(d2.to_string(index=False))
all_stats["2_CPI_AllItems_vs_Housing"] = d2

# Correlation: do housing costs inflate faster than the general index?
pivot_cpi = cpi_broad.pivot_table(
    index="Month_dt", columns="Commodity_Group", values="YoY_Inflation_Pct"
)
corr_cpi = pivot_cpi.corr().round(3)
print("\n[Correlation between series]")
print(corr_cpi.to_string())
all_stats["2_CPI_Correlation"] = corr_cpi.reset_index()

# ══════════════════════════════════════════════════════════════════════════════
# 3. INFLATION — Inflation by Income Decile (monthly, 2017–2023)
#    Key for energy poverty: do lower deciles face higher inflation?
# ══════════════════════════════════════════════════════════════════════════════
section("3. INFLATION BY INCOME DECILE (Monthly, 2017–2023)")
decile_inf = pd.read_csv(DATA_DIR / "Inflation_by_Income_Decile_2017_2023.csv")

d3 = desc(decile_inf, "YoY_Inflation_Pct",
          group_cols=["Income_Decile"])
print(d3.to_string(index=False))
all_stats["3_Inflation_by_Decile"] = d3

# Decile spread: max minus min inflation at each time point
decile_inf["Month_dt"] = pd.to_datetime(decile_inf["Month_dt"])
no_all = decile_inf[decile_inf["Income_Decile"] != "All deciles"]
spread = (
    no_all.groupby("Month_dt")["YoY_Inflation_Pct"]
    .agg(Spread=lambda x: x.max() - x.min())
)
d3_spread = desc(spread.reset_index(), "Spread",
                 label="Inflation Spread (max decile − min decile, pp)")
print("\n[Inflation spread across deciles]")
print(d3_spread.to_string(index=False))
all_stats["3_Inflation_Decile_Spread"] = d3_spread

# ══════════════════════════════════════════════════════════════════════════════
# 4. ENERGY PRICES — CPI for Energy Products (annual, 2007–2023)
#    Index values for electricity, gas, petrol, diesel, solid fuel
# ══════════════════════════════════════════════════════════════════════════════
section("4. ENERGY CPI BY PRODUCT (Annual, 2007–2023)")
energy_cpi = pd.read_csv(DATA_DIR / "CPI_Energy_Products_2007_2023.csv")

d4 = desc(energy_cpi, "CPI_Index",
          group_cols=["Energy_Product"])
print(d4.to_string(index=False))
all_stats["4_Energy_CPI_by_Product"] = d4

# YoY % change per product
energy_cpi_wide = energy_cpi.pivot(
    index="Year", columns="Energy_Product", values="CPI_Index"
)
energy_yoy = energy_cpi_wide.pct_change() * 100
energy_yoy_long = energy_yoy.reset_index().melt(
    id_vars="Year", var_name="Energy_Product", value_name="YoY_Pct_Change"
).dropna()

d4b = desc(energy_yoy_long, "YoY_Pct_Change",
           group_cols=["Energy_Product"])
print("\n[YoY % Change by Product]")
print(d4b.to_string(index=False))
all_stats["4_Energy_CPI_YoY_by_Product"] = d4b

# ══════════════════════════════════════════════════════════════════════════════
# 5. ENERGY PRICES — Residential Gas & Electricity (semi-annual, 2015–2023)
# ══════════════════════════════════════════════════════════════════════════════
section("5. RESIDENTIAL GAS & ELECTRICITY PRICES (Semi-Annual, 2015–2023)")

gas  = pd.read_csv(DATA_DIR / "Gas_Prices_Residential_2015_2023.csv")
elec = pd.read_csv(DATA_DIR / "Electricity_Prices_Residential_2015_2023.csv")

d5a = desc(gas,  "Price_EUR_per_GJ",  label="Gas Price (€/GJ, all-in)")
d5b = desc(elec, "Price_EUR_per_kWh", label="Electricity Price (€/kWh, all-in)")
d5 = pd.concat([d5a, d5b], ignore_index=True)
print(d5.to_string(index=False))

# By year
d5a_yr = gas.groupby("Year")["Price_EUR_per_GJ"].mean().round(4).reset_index()
d5b_yr = elec.groupby("Year")["Price_EUR_per_kWh"].mean().round(4).reset_index()
print("\n[Gas — Annual Average Price €/GJ]")
print(d5a_yr.to_string(index=False))
print("\n[Electricity — Annual Average Price €/kWh]")
print(d5b_yr.to_string(index=False))

all_stats["5_Gas_Elec_Prices_Overall"]    = d5
all_stats["5_Gas_Annual_Avg"]             = d5a_yr
all_stats["5_Electricity_Annual_Avg"]     = d5b_yr

# ══════════════════════════════════════════════════════════════════════════════
# 6. INCOME — Median Real Disposable Income by Age Group (annual, 2007–2023)
# ══════════════════════════════════════════════════════════════════════════════
section("6. INCOME — Median Real Disposable Income by Age Group (Annual, 2007–2023)")
income = pd.read_csv(DATA_DIR / "Income_Median_by_AgeGroup_2007_2023.csv")

d6 = desc(income, "Median_Real_Disposable_Income_EUR",
          group_cols=["Age_Group"])
print(d6.to_string(index=False))
all_stats["6_Income_by_AgeGroup"] = d6

# YoY % change (income volatility)
income_vol = pd.read_csv(DATA_DIR / "Income_Volatility_YoY_by_AgeGroup_2007_2023.csv")
vol_long = income_vol.melt(id_vars="Year", var_name="Age_Group",
                           value_name="YoY_Pct_Change").dropna()
d6b = desc(vol_long, "YoY_Pct_Change", group_cols=["Age_Group"])
print("\n[Income Volatility — YoY % Change]")
print(d6b.to_string(index=False))
all_stats["6_Income_Volatility"] = d6b

# ══════════════════════════════════════════════════════════════════════════════
# 7. INCOME — Disposable Income by Tenure (annual, 2007–2023)
# ══════════════════════════════════════════════════════════════════════════════
section("7. INCOME BY TENURE TYPE (Annual, 2007–2023)")
tenure = pd.read_csv(DATA_DIR / "Disposable_Income_by_Tenure_2007_2023.csv")

d7 = desc(tenure, "Median_Real_Disposable_Income_EUR",
          group_cols=["Tenure_Type"])
print(d7.to_string(index=False))

# Income gap: owner vs renter
tenure_wide = tenure.pivot(index="Year", columns="Tenure_Type",
                            values="Median_Real_Disposable_Income_EUR")
tenure_wide["Income_Gap_EUR"] = (
    tenure_wide["Owner-occupied"] - tenure_wide["Rented or rent free"]
)
print("\n[Owner vs Renter Income Gap (€)]")
print(tenure_wide[["Income_Gap_EUR"]].round(0).to_string())
all_stats["7_Income_by_Tenure"]   = d7
all_stats["7_Tenure_Income_Gap"]  = tenure_wide.reset_index()

# ══════════════════════════════════════════════════════════════════════════════
# 8. EMPLOYMENT / UNEMPLOYMENT (Quarterly ILO 2007–2017 + Monthly 2007–2023)
# ══════════════════════════════════════════════════════════════════════════════
section("8. EMPLOYMENT & UNEMPLOYMENT RATES (2007–2023)")

emp_q = pd.read_csv(DATA_DIR / "Employment_Rates_Quarterly_ILO_2007_2017.csv")
unemp = pd.read_csv(DATA_DIR / "Unemployment_Rate_Monthly_2007_2023.csv")

d8a = desc(emp_q, "Rate_Pct", group_cols=["Statistic"])
print("[Quarterly ILO Rates 2007–2017]")
print(d8a.to_string(index=False))

d8b = desc(unemp, "Unemployment_Rate_Pct", label="Monthly Unemployment Rate % (SA, 2007–2023)")
print("\n[Monthly Seasonally Adjusted Unemployment Rate 2007–2023]")
print(d8b.to_string(index=False))

unemp["Year"] = unemp["Month"].str[:4].astype(int)
d8b_yr = (
    unemp.groupby("Year")["Unemployment_Rate_Pct"]
    .mean()
    .round(2)
    .reset_index()
    .rename(columns={"Unemployment_Rate_Pct": "Avg_Unemployment_Rate_Pct"})
)
print("\n[Annual Average Unemployment Rate]")
print(d8b_yr.to_string(index=False))

all_stats["8_Employment_ILO_Quarterly"]  = d8a
all_stats["8_Unemployment_Overall"]      = d8b
all_stats["8_Unemployment_Annual_Avg"]   = d8b_yr

# ══════════════════════════════════════════════════════════════════════════════
# 9. POVERTY RATES (annual, 2007–2023)
#    Proxy for energy poverty transitions
# ══════════════════════════════════════════════════════════════════════════════
section("9. POVERTY RATES (Annual, 2007–2023)")
poverty = pd.read_csv(DATA_DIR / "Poverty_Rates_2007_2023.csv")

d9 = desc(poverty, "VALUE", group_cols=["Statistic"])
print(d9.to_string(index=False))

# Trend: poverty rate by year and statistic
print("\n[Poverty Rates by Year]")
print(poverty.pivot_table(
    index="Year", columns="Statistic", values="VALUE"
).round(2).to_string())

all_stats["9_Poverty_Rates"] = d9
all_stats["9_Poverty_by_Year"] = poverty.pivot_table(
    index="Year", columns="Statistic", values="VALUE"
).round(2).reset_index()

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY TABLE — Key Variables for Energy Poverty Transition Analysis
# ══════════════════════════════════════════════════════════════════════════════
section("SUMMARY — Key Variables (Energy Poverty Transition Analysis)")

summary_rows = [
    ("Housing CPI – YoY Inflation (%)",
     housing_cpi["YoY_Pct_Change"]),
    ("Housing CPI – Index Level",
     housing_cpi["CPI_Index"]),
    ("All-Items CPI – YoY Inflation (%)",
     cpi_broad.loc[cpi_broad["Commodity_Group"] == "All items", "YoY_Inflation_Pct"]),
    ("Energy CPI – Electricity (index)",
     energy_cpi.loc[energy_cpi["Energy_Product"] == "Electricity", "CPI_Index"]),
    ("Energy CPI – Natural Gas (index)",
     energy_cpi.loc[energy_cpi["Energy_Product"] == "Natural gas", "CPI_Index"]),
    ("Energy CPI – Electricity YoY (%)",
     energy_yoy_long.loc[energy_yoy_long["Energy_Product"] == "Electricity", "YoY_Pct_Change"]),
    ("Energy CPI – Natural Gas YoY (%)",
     energy_yoy_long.loc[energy_yoy_long["Energy_Product"] == "Natural gas", "YoY_Pct_Change"]),
    ("Energy CPI – Solid Fuel YoY (%)",
     energy_yoy_long.loc[energy_yoy_long["Energy_Product"].str.startswith("Solid"), "YoY_Pct_Change"]),
    ("Inflation – Bottom Decile (1st) YoY (%)",
     decile_inf.loc[decile_inf["Income_Decile"] == "1st decile", "YoY_Inflation_Pct"]),
    ("Inflation – Top Decile (10th) YoY (%)",
     decile_inf.loc[decile_inf["Income_Decile"] == "10th decile", "YoY_Inflation_Pct"]),
    ("Inflation Spread Across Deciles (pp)",
     spread["Spread"]),
    ("Residential Gas Price (€/GJ)",
     gas["Price_EUR_per_GJ"]),
    ("Residential Electricity Price (€/kWh)",
     elec["Price_EUR_per_kWh"]),
    ("Median Real Disposable Income 18–64 (€)",
     income.loc[income["Age_Group"] == "18 - 64 years", "Median_Real_Disposable_Income_EUR"]),
    ("Median Real Disposable Income 65+ (€)",
     income.loc[income["Age_Group"] == "65 years and over", "Median_Real_Disposable_Income_EUR"]),
    ("Income YoY Volatility 18–64 (%)",
     vol_long.loc[vol_long["Age_Group"].str.contains("18"), "YoY_Pct_Change"]),
    ("Income YoY Volatility 65+ (%)",
     vol_long.loc[vol_long["Age_Group"].str.contains("65"), "YoY_Pct_Change"]),
    ("Owner-Occupied Median Income (€)",
     tenure.loc[tenure["Tenure_Type"] == "Owner-occupied", "Median_Real_Disposable_Income_EUR"]),
    ("Rented Median Income (€)",
     tenure.loc[tenure["Tenure_Type"] == "Rented or rent free", "Median_Real_Disposable_Income_EUR"]),
    ("ILO Unemployment Rate – Quarterly (%)",
     emp_q.loc[emp_q["Statistic"].str.contains("Unemployment Rates.*74", na=False), "Rate_Pct"]),
    ("Unemployment Rate – Monthly SA (%)",
     unemp["Unemployment_Rate_Pct"]),
]

summary_records = []
for label, s in summary_rows:
    row = _full_stats(s.reset_index(drop=True), label)
    row.pop("Variable", None)
    summary_records.append({"Variable": label, **row})

summary = pd.DataFrame(summary_records)

# Print a compact view (central tendency + dispersion + shape only)
compact_cols = [
    "Variable", "N",
    "N_Positive", "N_Negative", "Pct_Positive", "Pct_Negative",
    "Mean", "Trimmed_Mean_5", "Median", "Mode",
    "Std_Dev", "Variance", "CV_Pct", "Range", "IQR", "MAD",
    "Min", "P5", "P10", "P25", "P75", "P90", "P95", "Max",
    "Skewness", "Kurtosis_Excess",
]
print(summary[[c for c in compact_cols if c in summary.columns]].to_string(index=False))
all_stats["SUMMARY"] = summary

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT — Single sheet with all variables, separated by blank rows + headers
# ══════════════════════════════════════════════════════════════════════════════
out_path = OUT_DIR / "Descriptive_Statistics.xlsx"

section_labels = {
    "1_Housing_CPI_Overall":         "1. Housing CPI — Overall (Monthly, 2007–2023)",
    "1_Housing_CPI_Annual_Avg":      "1. Housing CPI — Annual Averages",
    "2_CPI_AllItems_vs_Housing":     "2. All Items vs Housing CPI YoY % (Monthly, 2007–2023)",
    "2_CPI_Correlation":             "2. Correlation: All Items vs Housing CPI",
    "3_Inflation_by_Decile":         "3. Inflation by Income Decile (Monthly, 2017–2023)",
    "3_Inflation_Decile_Spread":     "3. Inflation Spread Across Deciles",
    "4_Energy_CPI_by_Product":       "4. Energy CPI by Product — Index Level (Annual, 2007–2023)",
    "4_Energy_CPI_YoY_by_Product":   "4. Energy CPI by Product — YoY % Change",
    "5_Gas_Elec_Prices_Overall":     "5. Residential Gas & Electricity Prices — Overall (2015–2023)",
    "5_Gas_Annual_Avg":              "5. Gas Prices — Annual Average (€/GJ)",
    "5_Electricity_Annual_Avg":      "5. Electricity Prices — Annual Average (€/kWh)",
    "6_Income_by_AgeGroup":          "6. Median Real Disposable Income by Age Group (Annual, 2007–2023)",
    "6_Income_Volatility":           "6. Income Volatility — YoY % Change by Age Group",
    "7_Income_by_Tenure":            "7. Median Real Disposable Income by Tenure (Annual, 2007–2023)",
    "7_Tenure_Income_Gap":           "7. Owner vs Renter Income Gap (€)",
    "8_Employment_ILO_Quarterly":    "8. Employment & Participation Rates — Quarterly ILO (2007–2017)",
    "8_Unemployment_Overall":        "8. Unemployment Rate — Monthly SA Overall (2007–2023)",
    "8_Unemployment_Annual_Avg":     "8. Unemployment Rate — Annual Average",
    "9_Poverty_Rates":               "9. Poverty Rates — Overall",
    "9_Poverty_by_Year":             "9. Poverty Rates — By Year",
    "SUMMARY":                       "SUMMARY — All Key Variables (Energy Poverty Transition Analysis)",
}

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    # Write SUMMARY first on its own sheet for quick reference
    summary.to_excel(writer, sheet_name="Summary", index=False)

    # Write everything onto one "All Variables" sheet
    ws_name = "All Variables"
    current_row = 1
    workbook  = writer.book
    worksheet = workbook.create_sheet(ws_name)

    header_fill   = PatternFill("solid", fgColor="1F4E79")   # dark blue
    section_fill  = PatternFill("solid", fgColor="D6E4F0")   # light blue
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    section_font  = Font(bold=True, color="1F4E79", size=11)

    for key, df in all_stats.items():
        if key == "SUMMARY":
            continue  # already on its own sheet

        label = section_labels.get(key, key)

        # Section header row
        worksheet.cell(row=current_row, column=1, value=label)
        worksheet.cell(row=current_row, column=1).font  = section_font
        worksheet.cell(row=current_row, column=1).fill  = section_fill
        worksheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
        # Merge across columns
        worksheet.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=max(len(df.columns), 8)
        )
        current_row += 1

        # Column headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=current_row, column=col_idx, value=col_name)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Data rows
        for _, row_data in df.iterrows():
            for col_idx, val in enumerate(row_data, start=1):
                worksheet.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        # Blank separator row
        current_row += 2

    # Auto-fit column widths
    for col_cells in worksheet.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10
        )
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = min(max_len + 4, 50)

print(f"\n✅  Descriptive statistics exported → {out_path}")
print(f"    Sheets: 'Summary' (key variables) + 'All Variables' (full detail)")
